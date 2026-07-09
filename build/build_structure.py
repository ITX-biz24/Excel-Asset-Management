# -*- coding: utf-8 -*-
"""
構造ビルダ（openpyxl）
======================
個人資産管理システム v2 のブック構造（シート・数式・入力規則・条件付き書式・
グラフ・名前付き範囲・テーマ）を生成し、中間 .xlsx として保存する。

VBA は本スクリプトでは注入しない（openpyxl は VBA 非対応）。
VBA 注入と .xlsm 化は build/inject_vba.py が担当する。

設計は docs/SPEC.md を参照。口座はマスタ表で可変、残高は 2 本の SUMIFS で一般化する。
"""

import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import theme as T

# ------------------------------------------------------------
# パス
# ------------------------------------------------------------
BUILD_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR = os.path.dirname(BUILD_DIR)
OUT_XLSX = os.path.join(BUILD_DIR, "_structure.xlsx")

# ------------------------------------------------------------
# シート名（modConfig.bas と一致させること）
# ------------------------------------------------------------
SH = {
    "dash":  "Dashboard",
    "txn":   "取引履歴",
    "fixed": "固定収支",
    "sub":   "サブスク",
    "report":"レポート",
    "snap":  "スナップショット",
    "master":"マスタ",
    "config":"設定",
}

# テーブル displayName（ASCII で統一。列見出しは日本語）
TB = {
    "account":  "T_Account",
    "category": "T_Category",
    "pay":      "T_Pay",
    "txn":      "T_Txn",
    "fixed":    "T_Fixed",
    "sub":      "T_Sub",
    "snap":     "T_Snap",
}

# 既定の対象年月（今日 2026-07 に合わせる）
TARGET_YM = "2026-07"

# 取引テーブルの確保行数（テーブルは入力で自動拡張するが余裕を持って確保）
TXN_ROWS = 400

# マスタ初期値（カテゴリ・支払方法はここが唯一の定義。build_master と report で共用）
CATEGORIES = [
    ("給与", "収入"), ("賞与", "収入"), ("副業", "収入"), ("その他収入", "収入"),
    ("食費", "支出"), ("日用品", "支出"), ("住居", "支出"), ("水道光熱", "支出"),
    ("通信", "支出"), ("交通", "支出"), ("娯楽", "支出"), ("交際", "支出"),
    ("医療", "支出"), ("教育", "支出"), ("保険", "支出"), ("サブスク", "支出"),
    ("その他支出", "支出"),
]
PAYS = ["現金", "クレジット", "銀行振込", "口座引落", "QR決済", "デビット", "電子マネー"]

# データテーブル用のダーク系スタイル（明色テーブルより地色に馴染む）
TABLE_STYLE_DARK = "TableStyleDark9"


# ============================================================
# 汎用ヘルパ
# ============================================================
def paint(ws, max_row, max_col, color="bg"):
    """指定範囲をテーマ背景色で塗り、アプリらしい地色にする。"""
    f = T.fill(color)
    fnt = T.font()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = f
            cell.font = fnt


def put(ws, coord, value=None, *, font_kw=None, fill_c=None, align=None,
        fmt=None, border=None):
    """1セルへ値とスタイルをまとめて設定する小ヘルパ。"""
    cell = ws[coord]
    if value is not None:
        cell.value = value
    if font_kw is not None:
        cell.font = T.font(**font_kw)
    if fill_c is not None:
        cell.fill = T.fill(fill_c)
    if align is not None:
        cell.alignment = align
    if fmt is not None:
        cell.number_format = fmt
    if border is not None:
        cell.border = border
    return cell


def title_block(ws, text, sub=""):
    """各シート上部の見出し（アイコン＋タイトル＋サブ）を描く。"""
    put(ws, "B2", text, font_kw=dict(size=20, bold=True, color="text"),
        align=T.LEFT)
    if sub:
        put(ws, "B3", sub, font_kw=dict(size=10, color="subtext"), align=T.LEFT)


def add_table(ws, name, ref, style=T.TABLE_STYLE):
    """テーブル（ListObject）を追加する。"""
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)
    return tab


def add_name(wb, name, ref):
    """名前付き範囲を追加（テーブル列参照や絶対参照を受け付ける）。"""
    wb.defined_names.add(DefinedName(name=name, attr_text=ref))


def set_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def hide_grid(ws):
    ws.sheet_view.showGridLines = False


# ============================================================
# マスタシート
# ============================================================
def build_master(wb):
    ws = wb.create_sheet(SH["master"])
    hide_grid(ws)
    paint(ws, 60, 18)
    ws.sheet_properties.tabColor = T.COL["muted"]
    title_block(ws, "マスタ", "口座・カテゴリ・支払方法・種類の一元管理（ここを編集すれば全シートに反映）")
    set_widths(ws, {"A": 2, "B": 16, "C": 14, "D": 10, "E": 16,
                    "F": 2, "G": 16, "H": 12, "I": 2, "J": 16})

    # --- 口座テーブル（残高計算の中核。行を足すだけで口座を拡張できる）---
    hdr_row = 5
    acc_headers = ["口座名", "初期残高", "表示順", "現在残高"]
    for i, h in enumerate(acc_headers):
        put(ws, f"{get_column_letter(2+i)}{hdr_row}", h,
            font_kw=dict(bold=True, color="bg"), fill_c="accent", align=T.CENTER)
    accounts = [("現金", 100, 1), ("銀行", 100, 2), ("QR", 100, 3)]
    for j, (nm, init, order) in enumerate(accounts):
        r = hdr_row + 1 + j
        put(ws, f"B{r}", nm, fill_c="input", align=T.LEFT)
        put(ws, f"C{r}", init, fill_c="input", align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"D{r}", order, fill_c="input", align=T.CENTER)
        # 現在残高 = 初期残高 + 入金先合計 - 出金元合計（口座に依存しない一般式）
        # 行内参照は openpyxl では [@..] が壊れるため明示セル参照を用いる（列全体参照は構造化のまま）
        put(ws, f"E{r}",
            f"=C{r}+SUMIFS({TB['txn']}[金額],{TB['txn']}[入金先],B{r})"
            f"-SUMIFS({TB['txn']}[金額],{TB['txn']}[出金元],B{r})",
            align=T.RIGHT, fmt=T.FMT_YEN)
    add_table(ws, TB["account"], f"B{hdr_row}:E{hdr_row+len(accounts)}")
    put(ws, f"B{hdr_row-1}", "■ 口座", font_kw=dict(bold=True, color="teal"), align=T.LEFT)

    # 残高プリセット（強制セット）: 初期残高を実残高として強制適用するボタン領域
    btn_row = hdr_row + len(accounts) + 2   # 口座テーブルの下
    button_cell(ws, f"B{btn_row}", "▶ 残高を強制セット", color="peach")
    put(ws, f"B{btn_row+1}",
        "「初期残高」＝プリセット残高。ボタンで取引履歴を消去し、現在残高を初期残高に強制リセット（履歴参照なし）。",
        font_kw=dict(size=9, color="subtext"), align=T.LEFT)

    # --- カテゴリテーブル（区分で収入/支出を分類）---
    cat_hdr = 5
    put(ws, f"G{cat_hdr-1}", "■ カテゴリ", font_kw=dict(bold=True, color="teal"), align=T.LEFT)
    put(ws, f"G{cat_hdr}", "カテゴリ", font_kw=dict(bold=True, color="bg"),
        fill_c="accent", align=T.CENTER)
    put(ws, f"H{cat_hdr}", "区分", font_kw=dict(bold=True, color="bg"),
        fill_c="accent", align=T.CENTER)
    categories = CATEGORIES
    for j, (nm, kind) in enumerate(categories):
        r = cat_hdr + 1 + j
        put(ws, f"G{r}", nm, fill_c="input", align=T.LEFT)
        put(ws, f"H{r}", kind, fill_c="input", align=T.CENTER)
    add_table(ws, TB["category"], f"G{cat_hdr}:H{cat_hdr+len(categories)}")

    # --- 支払方法テーブル ---
    pay_hdr = 5
    put(ws, f"J{pay_hdr-1}", "■ 支払方法", font_kw=dict(bold=True, color="teal"), align=T.LEFT)
    put(ws, f"J{pay_hdr}", "支払方法", font_kw=dict(bold=True, color="bg"),
        fill_c="accent", align=T.CENTER)
    pays = PAYS
    for j, nm in enumerate(pays):
        put(ws, f"J{pay_hdr+1+j}", nm, fill_c="input", align=T.LEFT)
    add_table(ws, TB["pay"], f"J{pay_hdr}:J{pay_hdr+len(pays)}")

    # --- 種類リスト（収入/支出/振替）: 静的名前付き範囲 ---
    kinds_col = "L"
    put(ws, f"{kinds_col}{pay_hdr-1}", "■ 種類", font_kw=dict(bold=True, color="teal"), align=T.LEFT)
    kinds = ["収入", "支出", "振替"]
    for j, k in enumerate(kinds):
        put(ws, f"{kinds_col}{pay_hdr+j}", k, fill_c="card2", align=T.CENTER)
    ws.column_dimensions[kinds_col].width = 10

    # 名前付き範囲（テーブル列参照は入力規則の元として利用）
    add_name(wb, "口座リスト",   f"{SH['master']}!{TB['account']}[口座名]")
    add_name(wb, "カテゴリリスト", f"{SH['master']}!{TB['category']}[カテゴリ]")
    add_name(wb, "支払方法リスト", f"{SH['master']}!{TB['pay']}[支払方法]")
    add_name(wb, "種類リスト",
             f"{SH['master']}!${kinds_col}${pay_hdr}:${kinds_col}${pay_hdr+len(kinds)-1}")
    return ws


# ============================================================
# 取引履歴シート（全集計の唯一の真実）
# ============================================================
# 列定義: (見出し, 列文字, 幅, 数値書式, 計算列か)
TXN_COLS = [
    ("日付",     "B", 12, T.FMT_DATE, False),
    ("種類",     "C", 8,  None,       False),
    ("出金元",   "D", 10, None,       False),
    ("入金先",   "E", 10, None,       False),
    ("金額",     "F", 13, T.FMT_YEN0, False),
    ("カテゴリ", "G", 11, None,       False),
    ("支払方法", "H", 11, None,       False),
    ("メモ",     "I", 26, None,       False),
    ("年月",     "J", 9,  T.FMT_YM,   True),
]

# デモデータ（利用者が削除して使い始める想定。2026-07）
TXN_DEMO = [
    # 日付,       種類,   出金元, 入金先, 金額,   カテゴリ,     支払方法,   メモ
    ("2026-07-05", "収入", "",     "銀行", 250000, "給与",       "銀行振込", "7月給与"),
    ("2026-07-02", "収入", "",     "QR",   8000,   "副業",       "QR決済",   "ポイント収入"),
    ("2026-07-01", "支出", "銀行", "",     80000,  "住居",       "口座引落", "家賃"),
    ("2026-07-03", "支出", "現金", "",     3200,   "食費",       "現金",     "スーパー"),
    ("2026-07-04", "支出", "QR",   "",     1500,   "食費",       "QR決済",   "コンビニ"),
    ("2026-07-06", "支出", "銀行", "",     6000,   "通信",       "口座引落", "携帯料金"),
    ("2026-07-06", "支出", "QR",   "",     4800,   "娯楽",       "QR決済",   "映画"),
    ("2026-07-08", "支出", "現金", "",     2500,   "食費",       "現金",     "ランチ"),
    ("2026-07-08", "支出", "銀行", "",     12000,  "交際",       "クレジット", "飲み会"),
    ("2026-07-07", "振替", "銀行", "現金", 30000,  "",           "銀行振込", "ATM出金"),
]


def build_txn(wb):
    ws = wb.create_sheet(SH["txn"])
    hide_grid(ws)
    paint(ws, TXN_ROWS + 10, 12)
    ws.sheet_properties.tabColor = T.COL["accent"]
    title_block(ws, "取引履歴", "すべての残高・集計・グラフの元データ。1行=1取引")
    set_widths(ws, {"A": 2})
    for name, col, w, fmt, _ in TXN_COLS:
        ws.column_dimensions[col].width = w

    hdr = 6
    # ヘッダ
    for name, col, w, fmt, _ in TXN_COLS:
        put(ws, f"{col}{hdr}", name, font_kw=dict(bold=True, color="text"), align=T.CENTER)
    # データ行（デモ＋空行）に数式・書式を敷設
    first, last = hdr + 1, hdr + TXN_ROWS
    for r in range(first, last + 1):
        idx = r - first
        demo = TXN_DEMO[idx] if idx < len(TXN_DEMO) else None
        for ci, (name, col, w, fmt, calc) in enumerate(TXN_COLS):
            cell = ws[f"{col}{r}"]
            if calc:  # 年月（計算列。行内参照は明示セル参照で堅牢化）
                cell.value = f'=IF($B{r}="","",TEXT($B{r},"yyyy-mm"))'
                cell.fill = T.fill("card")  # 計算列は控えめな面で自動生成を示す
                cell.font = T.font(color="subtext")
                cell.alignment = T.CENTER
            elif demo is not None:
                v = demo[ci]
                if name == "日付" and v:
                    from datetime import datetime
                    cell.value = datetime.strptime(v, "%Y-%m-%d")
                else:
                    cell.value = v if v != "" else None
            if fmt:
                cell.number_format = fmt
            if name in ("金額",):
                cell.alignment = T.RIGHT
            elif name in ("種類", "出金元", "入金先", "支払方法"):
                cell.alignment = T.CENTER
            elif name in ("日付",):
                cell.alignment = T.CENTER

    ref = f"B{hdr}:J{last}"
    add_table(ws, TB["txn"], ref, style=TABLE_STYLE_DARK)
    ws.freeze_panes = f"B{hdr+1}"

    # --- 入力規則（一次防御。詳細な整合性チェックは VBA が担う）---
    data_rng = f"C{first}:C{last}"
    dv_kind = DataValidation(type="list", formula1="種類リスト", allow_blank=True,
                             showErrorMessage=True, errorTitle="種類",
                             error="収入 / 支出 / 振替 から選択してください")
    dv_kind.add(data_rng); ws.add_data_validation(dv_kind)

    dv_from = DataValidation(type="list", formula1="口座リスト", allow_blank=True)
    dv_from.add(f"D{first}:D{last}"); ws.add_data_validation(dv_from)
    dv_to = DataValidation(type="list", formula1="口座リスト", allow_blank=True)
    dv_to.add(f"E{first}:E{last}"); ws.add_data_validation(dv_to)

    dv_amt = DataValidation(type="decimal", operator="greaterThan", formula1="0",
                            allow_blank=True, showErrorMessage=True,
                            errorTitle="金額", error="金額は0より大きい値を入力してください")
    dv_amt.add(f"F{first}:F{last}"); ws.add_data_validation(dv_amt)

    dv_cat = DataValidation(type="list", formula1="カテゴリリスト", allow_blank=True)
    dv_cat.add(f"G{first}:G{last}"); ws.add_data_validation(dv_cat)
    dv_pay = DataValidation(type="list", formula1="支払方法リスト", allow_blank=True)
    dv_pay.add(f"H{first}:H{last}"); ws.add_data_validation(dv_pay)

    dv_date = DataValidation(type="date", operator="greaterThanOrEqual",
                             formula1="DATE(2000,1,1)", allow_blank=True,
                             showErrorMessage=True, errorTitle="日付",
                             error="正しい日付を入力してください")
    dv_date.add(f"B{first}:B{last}"); ws.add_data_validation(dv_date)

    # --- 条件付き書式（種類で金額を色分け）---
    ws.conditional_formatting.add(
        f"C{first}:C{last}",
        FormulaRule(formula=[f'$C{first}="収入"'], font=Font(color=T.COL["green"], bold=True)))
    ws.conditional_formatting.add(
        f"C{first}:C{last}",
        FormulaRule(formula=[f'$C{first}="支出"'], font=Font(color=T.COL["red"], bold=True)))
    ws.conditional_formatting.add(
        f"C{first}:C{last}",
        FormulaRule(formula=[f'$C{first}="振替"'], font=Font(color=T.COL["teal"], bold=True)))
    ws.conditional_formatting.add(
        f"F{first}:F{last}",
        FormulaRule(formula=[f'$C{first}="収入"'], font=Font(color=T.COL["green"])))
    ws.conditional_formatting.add(
        f"F{first}:F{last}",
        FormulaRule(formula=[f'$C{first}="支出"'], font=Font(color=T.COL["red"])))
    return ws


# ============================================================
# ボタン風セル（COM で実ボタンを重ねる際のアンカー兼、視覚的アフォーダンス）
# ============================================================
def button_cell(ws, coord, label, color="accent"):
    c = ws[coord]
    c.value = label
    c.font = T.font(size=11, bold=True, color="bg")
    c.fill = T.fill(color)
    c.alignment = T.CENTER
    return c


def _list_dv(ws, formula1, rng, allow_blank=True, title=None, error=None):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank,
                        showErrorMessage=bool(error), errorTitle=title, error=error)
    dv.add(rng); ws.add_data_validation(dv)
    return dv


# ============================================================
# 固定収支シート
# ============================================================
FIXED_ROWS = 40
# (見出し, 列, 幅, 書式, 配置)
FIXED_COLS = [
    ("名称",     "B", 18, None,       T.LEFT),
    ("種類",     "C", 8,  None,       T.CENTER),
    ("金額",     "D", 13, T.FMT_YEN0, T.RIGHT),
    ("支払日",   "E", 8,  T.FMT_INT,  T.CENTER),
    ("口座",     "F", 10, None,       T.CENTER),
    ("カテゴリ", "G", 11, None,       T.CENTER),
    ("支払方法", "H", 11, None,       T.CENTER),
    ("有効",     "I", 7,  None,       T.CENTER),
    ("メモ",     "J", 22, None,       T.LEFT),
]
FIXED_DEMO = [
    ("給与",   "収入", 250000, 25, "銀行", "給与",     "銀行振込", "ON", "月給"),
    ("家賃",   "支出", 80000,  27, "銀行", "住居",     "口座引落", "ON", ""),
    ("電気代", "支出", 6000,   15, "銀行", "水道光熱", "口座引落", "ON", ""),
    ("携帯代", "支出", 6000,   10, "銀行", "通信",     "口座引落", "ON", ""),
    ("保険",   "支出", 5000,   5,  "銀行", "保険",     "口座引落", "OFF", "見直し中"),
]


def build_fixed(wb):
    ws = wb.create_sheet(SH["fixed"])
    hide_grid(ws)
    paint(ws, FIXED_ROWS + 12, 12)
    ws.sheet_properties.tabColor = T.COL["mauve"]
    title_block(ws, "固定収支", "毎月の定期的な収入・支出。ボタンで当月の取引履歴へ一括反映")
    set_widths(ws, {"A": 2})
    for name, col, w, *_ in FIXED_COLS:
        ws.column_dimensions[col].width = w

    # 反映ボタン（COMで実ボタンを重ねる）と当月合計
    button_cell(ws, "B4", "▶ 今月へ反映", color="green")
    put(ws, "D4", "固定支出/月:", font_kw=dict(color="subtext"), align=T.RIGHT)
    put(ws, "E4", '=SUMIFS(T_Fixed[金額],T_Fixed[種類],"支出",T_Fixed[有効],"ON")',
        font_kw=dict(bold=True, color="red"), align=T.LEFT, fmt=T.FMT_YEN0)
    put(ws, "G4", "固定収入/月:", font_kw=dict(color="subtext"), align=T.RIGHT)
    put(ws, "H4", '=SUMIFS(T_Fixed[金額],T_Fixed[種類],"収入",T_Fixed[有効],"ON")',
        font_kw=dict(bold=True, color="green"), align=T.LEFT, fmt=T.FMT_YEN0)

    hdr = 6
    for name, col, w, *_ in FIXED_COLS:
        put(ws, f"{col}{hdr}", name, font_kw=dict(bold=True, color="text"), align=T.CENTER)
    first, last = hdr + 1, hdr + FIXED_ROWS
    for r in range(first, last + 1):
        idx = r - first
        demo = FIXED_DEMO[idx] if idx < len(FIXED_DEMO) else None
        for ci, (name, col, w, fmt, align) in enumerate(FIXED_COLS):
            cell = ws[f"{col}{r}"]
            if demo is not None:
                v = demo[ci]
                cell.value = v if v != "" else None
            if fmt:
                cell.number_format = fmt
            cell.alignment = align
    add_table(ws, TB["fixed"], f"B{hdr}:J{last}", style=TABLE_STYLE_DARK)
    ws.freeze_panes = f"B{hdr+1}"

    _list_dv(ws, "種類リスト", f"C{first}:C{last}")
    _list_dv(ws, "口座リスト", f"F{first}:F{last}")
    _list_dv(ws, "カテゴリリスト", f"G{first}:G{last}")
    _list_dv(ws, "支払方法リスト", f"H{first}:H{last}")
    _list_dv(ws, '"ON,OFF"', f"I{first}:I{last}")
    dv_amt = DataValidation(type="decimal", operator="greaterThan", formula1="0",
                            allow_blank=True)
    dv_amt.add(f"D{first}:D{last}"); ws.add_data_validation(dv_amt)
    dv_day = DataValidation(type="whole", operator="between", formula1="1", formula2="31",
                            allow_blank=True, showErrorMessage=True, errorTitle="支払日",
                            error="1〜31の日を入力してください")
    dv_day.add(f"E{first}:E{last}"); ws.add_data_validation(dv_day)

    # 有効=OFF はグレーアウト
    ws.conditional_formatting.add(
        f"B{first}:J{last}",
        FormulaRule(formula=[f'$I{first}="OFF"'], font=Font(color=T.COL["muted"], italic=True)))
    return ws


# ============================================================
# サブスクシート
# ============================================================
SUB_ROWS = 40
SUB_COLS = [
    ("名称",     "B", 20, None,       T.LEFT),
    ("開始日",   "C", 12, T.FMT_DATE, T.CENTER),
    ("支払日",   "D", 8,  T.FMT_INT,  T.CENTER),
    ("金額",     "E", 12, T.FMT_YEN0, T.RIGHT),
    ("支払方法", "F", 11, None,       T.CENTER),
    ("口座",     "G", 10, None,       T.CENTER),
    ("カテゴリ", "H", 11, None,       T.CENTER),
    ("有効",     "I", 7,  None,       T.CENTER),
    ("メモ",     "J", 20, None,       T.LEFT),
]
SUB_DEMO = [
    ("Netflix", "2025-01-01", 15, 1490, "クレジット", "銀行", "サブスク", "ON", "動画"),
    ("Spotify", "2025-03-01", 5,  980,  "クレジット", "銀行", "サブスク", "ON", "音楽"),
    ("クラウド", "2026-01-01", 1,  500,  "QR決済",   "QR",   "サブスク", "OFF", "解約検討"),
]


def build_sub(wb):
    ws = wb.create_sheet(SH["sub"])
    hide_grid(ws)
    paint(ws, SUB_ROWS + 12, 12)
    ws.sheet_properties.tabColor = T.COL["peach"]
    title_block(ws, "サブスク", "定期課金の管理。有効なものを当月の取引履歴へ反映")
    set_widths(ws, {"A": 2})
    for name, col, w, *_ in SUB_COLS:
        ws.column_dimensions[col].width = w

    button_cell(ws, "B4", "▶ 今月へ反映", color="green")
    put(ws, "D4", "有効サブスク/月:", font_kw=dict(color="subtext"), align=T.RIGHT)
    put(ws, "E4", '=SUMIFS(T_Sub[金額],T_Sub[有効],"ON")',
        font_kw=dict(bold=True, color="peach"), align=T.LEFT, fmt=T.FMT_YEN0)
    put(ws, "G4", "契約数(有効):", font_kw=dict(color="subtext"), align=T.RIGHT)
    put(ws, "H4", '=COUNTIFS(T_Sub[有効],"ON")',
        font_kw=dict(bold=True, color="text"), align=T.LEFT, fmt=T.FMT_INT)

    hdr = 6
    for name, col, w, *_ in SUB_COLS:
        put(ws, f"{col}{hdr}", name, font_kw=dict(bold=True, color="text"), align=T.CENTER)
    first, last = hdr + 1, hdr + SUB_ROWS
    for r in range(first, last + 1):
        idx = r - first
        demo = SUB_DEMO[idx] if idx < len(SUB_DEMO) else None
        for ci, (name, col, w, fmt, align) in enumerate(SUB_COLS):
            cell = ws[f"{col}{r}"]
            if demo is not None:
                v = demo[ci]
                if name == "開始日" and v:
                    from datetime import datetime
                    cell.value = datetime.strptime(v, "%Y-%m-%d")
                else:
                    cell.value = v if v != "" else None
            if fmt:
                cell.number_format = fmt
            cell.alignment = align
    add_table(ws, TB["sub"], f"B{hdr}:J{last}", style=TABLE_STYLE_DARK)
    ws.freeze_panes = f"B{hdr+1}"

    _list_dv(ws, "支払方法リスト", f"F{first}:F{last}")
    _list_dv(ws, "口座リスト", f"G{first}:G{last}")
    _list_dv(ws, "カテゴリリスト", f"H{first}:H{last}")
    _list_dv(ws, '"ON,OFF"', f"I{first}:I{last}")
    dv_amt = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=True)
    dv_amt.add(f"E{first}:E{last}"); ws.add_data_validation(dv_amt)
    dv_day = DataValidation(type="whole", operator="between", formula1="1", formula2="31",
                            allow_blank=True)
    dv_day.add(f"D{first}:D{last}"); ws.add_data_validation(dv_day)

    ws.conditional_formatting.add(
        f"B{first}:J{last}",
        FormulaRule(formula=[f'$I{first}="OFF"'], font=Font(color=T.COL["muted"], italic=True)))
    return ws


# ============================================================
# 設定シート
# ============================================================
def build_config(wb):
    ws = wb.create_sheet(SH["config"])
    hide_grid(ws)
    paint(ws, 40, 12)
    ws.sheet_properties.tabColor = T.COL["muted"]
    title_block(ws, "設定", "システム全体の基準値。数式・VBA が参照する定数の置き場")
    set_widths(ws, {"A": 2, "B": 22, "C": 20, "D": 40})

    rows = [
        ("対象年月",       TARGET_YM, "ダッシュボード・当月集計の基準（yyyy-mm）"),
        ("集計開始年月",   "2026-01", "レポート月次サマリの開始月（yyyy-mm）"),
        ("残高不足チェック", "ON",      "出金時に残高不足を拒否するか（ON/OFF）"),
        ("自動バックアップ", "OFF",     "保存前に自動バックアップするか（ON/OFF）"),
        ("バックアップ先",  r"backup", "バックアップ保存フォルダ（相対 or 絶対パス）"),
        ("アプリ名",       "個人資産管理システム", ""),
        ("バージョン",     "v2.0",    ""),
    ]
    start = 5
    put(ws, f"B{start-1}", "■ 基本設定", font_kw=dict(bold=True, color="teal"), align=T.LEFT)
    for i, (k, v, note) in enumerate(rows):
        r = start + i
        put(ws, f"B{r}", k, font_kw=dict(bold=True, color="text"),
            fill_c="card", align=T.LEFT)
        put(ws, f"C{r}", v, fill_c="input", align=T.CENTER)
        put(ws, f"D{r}", note, font_kw=dict(size=9, color="subtext"), align=T.LEFT)

    # 名前付き範囲（数式から参照）
    add_name(wb, "対象年月",     f"{SH['config']}!$C${start}")
    add_name(wb, "集計開始年月", f"{SH['config']}!$C${start+1}")
    add_name(wb, "残高不足チェック", f"{SH['config']}!$C${start+2}")
    add_name(wb, "自動バックアップ", f"{SH['config']}!$C${start+3}")
    add_name(wb, "バックアップ先", f"{SH['config']}!$C${start+4}")
    return ws


# ============================================================
# スナップショットシート（資産推移の元データ）
# ============================================================
def build_snap(wb):
    ws = wb.create_sheet(SH["snap"])
    hide_grid(ws)
    paint(ws, 60, 10)
    ws.sheet_properties.tabColor = T.COL["teal"]
    title_block(ws, "スナップショット", "月末ごとの資産残高。VBA『月末保存』で追記され資産推移グラフの元になる")
    set_widths(ws, {"A": 2, "B": 11, "C": 14, "D": 12, "E": 12, "F": 12})

    hdr = 6
    heads = ["年月", "総資産", "現金", "銀行", "QR"]
    for i, h in enumerate(heads):
        put(ws, f"{get_column_letter(2+i)}{hdr}", h,
            font_kw=dict(bold=True, color="text"), align=T.CENTER)
    # 初期行: 当月のライブ値（VBA が翌月以降を値として追記していく）
    r = hdr + 1
    put(ws, f"B{r}", "=対象年月", align=T.CENTER)
    put(ws, f"C{r}", "=SUM(T_Account[現在残高])", align=T.RIGHT, fmt=T.FMT_YEN)
    put(ws, f"D{r}", '=IFERROR(INDEX(T_Account[現在残高],MATCH("現金",T_Account[口座名],0)),0)',
        align=T.RIGHT, fmt=T.FMT_YEN)
    put(ws, f"E{r}", '=IFERROR(INDEX(T_Account[現在残高],MATCH("銀行",T_Account[口座名],0)),0)',
        align=T.RIGHT, fmt=T.FMT_YEN)
    put(ws, f"F{r}", '=IFERROR(INDEX(T_Account[現在残高],MATCH("QR",T_Account[口座名],0)),0)',
        align=T.RIGHT, fmt=T.FMT_YEN)
    add_table(ws, TB["snap"], f"B{hdr}:F{r}", style=TABLE_STYLE_DARK)
    return ws


# ============================================================
# レポートシート（月次・年間・ランキング）
# ============================================================
def build_report(wb):
    ws = wb.create_sheet(SH["report"])
    hide_grid(ws)
    paint(ws, 60, 14)
    ws.sheet_properties.tabColor = T.COL["yellow"]
    title_block(ws, "レポート", "月次サマリ・年間履歴・カテゴリ/支払方法ランキング（グラフの元データ）")
    set_widths(ws, {"A": 2, "B": 11, "C": 12, "D": 12, "E": 12, "F": 9, "G": 13,
                    "H": 2, "I": 8, "J": 13, "K": 13, "L": 13})

    def section(coord, text):
        put(ws, coord, text, font_kw=dict(bold=True, color="teal"), align=T.LEFT)

    def header_row(row, cols, start_col=2):
        for i, h in enumerate(cols):
            put(ws, f"{get_column_letter(start_col+i)}{row}", h,
                font_kw=dict(bold=True, color="bg"), fill_c="accent", align=T.CENTER)

    # --- 月次サマリ（24か月）---
    section("B4", "■ 月次サマリ")
    mhdr = 5
    header_row(mhdr, ["年月", "収入", "支出", "収支", "貯蓄率", "月末総資産"])
    mfirst, mlast = mhdr + 1, mhdr + 24
    for r in range(mfirst, mlast + 1):
        off = r - mfirst
        put(ws, f"B{r}", f'=TEXT(EDATE(DATEVALUE(集計開始年月&"-01"),{off}),"yyyy-mm")',
            align=T.CENTER, fill_c="card")
        put(ws, f"C{r}", f'=SUMIFS(T_Txn[金額],T_Txn[種類],"収入",T_Txn[年月],$B{r})',
            align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"D{r}", f'=SUMIFS(T_Txn[金額],T_Txn[種類],"支出",T_Txn[年月],$B{r})',
            align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"E{r}", f"=C{r}-D{r}", align=T.RIGHT, fmt=T.FMT_YEN)
        put(ws, f"F{r}", f'=IFERROR(E{r}/C{r},"")', align=T.CENTER, fmt=T.FMT_PCT)
        put(ws, f"G{r}", f'=IFERROR(INDEX(T_Snap[総資産],MATCH($B{r},T_Snap[年月],0)),"")',
            align=T.RIGHT, fmt=T.FMT_YEN0)

    # --- 年間履歴（3年）---
    section("I4", "■ 年間履歴")
    yhdr = 5
    header_row(yhdr, ["年", "年間収入", "年間支出", "年間貯蓄"], start_col=9)
    yfirst, ylast = yhdr + 1, yhdr + 3
    for r in range(yfirst, ylast + 1):
        off = r - yfirst
        put(ws, f"I{r}", f"=VALUE(LEFT(集計開始年月,4))+{off}", align=T.CENTER, fill_c="card")
        put(ws, f"J{r}",
            f'=SUMPRODUCT((LEFT(T_Txn[年月],4)=$I{r}&"")*(T_Txn[種類]="収入")*T_Txn[金額])',
            align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"K{r}",
            f'=SUMPRODUCT((LEFT(T_Txn[年月],4)=$I{r}&"")*(T_Txn[種類]="支出")*T_Txn[金額])',
            align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"L{r}", f"=J{r}-K{r}", align=T.RIGHT, fmt=T.FMT_YEN)

    # --- カテゴリランキング ---
    ncat = len(CATEGORIES)
    section("B31", "■ カテゴリ別支出ランキング")
    chdr = 32
    header_row(chdr, ["カテゴリ", "当月支出", "当年支出", "順位"])
    cfirst, clast = chdr + 1, chdr + ncat
    for r in range(cfirst, clast + 1):
        idx = r - cfirst + 1
        put(ws, f"B{r}", f'=IFERROR(INDEX(T_Category[カテゴリ],{idx}),"")', align=T.LEFT, fill_c="card")
        put(ws, f"C{r}",
            f'=IF($B{r}="",0,SUMIFS(T_Txn[金額],T_Txn[種類],"支出",T_Txn[カテゴリ],$B{r},T_Txn[年月],対象年月))',
            align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"D{r}",
            f'=IF($B{r}="",0,SUMPRODUCT((LEFT(T_Txn[年月],4)=LEFT(対象年月,4))*(T_Txn[種類]="支出")*(T_Txn[カテゴリ]=$B{r})*T_Txn[金額]))',
            align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"E{r}", f'=IF(C{r}=0,"",RANK(C{r},$C${cfirst}:$C${clast}))',
            align=T.CENTER, fmt=T.FMT_INT)

    # --- 支払方法ランキング ---
    npay = len(PAYS)
    section("I31", "■ 支払方法別ランキング")
    phdr = 32
    header_row(phdr, ["支払方法", "当月支出", "順位"], start_col=9)
    pfirst, plast = phdr + 1, phdr + npay
    for r in range(pfirst, plast + 1):
        idx = r - pfirst + 1
        put(ws, f"I{r}", f'=IFERROR(INDEX(T_Pay[支払方法],{idx}),"")', align=T.LEFT, fill_c="card")
        put(ws, f"J{r}",
            f'=IF($I{r}="",0,SUMIFS(T_Txn[金額],T_Txn[種類],"支出",T_Txn[支払方法],$I{r},T_Txn[年月],対象年月))',
            align=T.RIGHT, fmt=T.FMT_YEN0)
        put(ws, f"K{r}", f'=IF(J{r}=0,"",RANK(J{r},$J${pfirst}:$J${plast}))',
            align=T.CENTER, fmt=T.FMT_INT)

    # チャート参照に使う範囲を属性として持たせる
    ws._rng = dict(month=(mfirst, mlast), cat=(cfirst, clast))
    return ws


# ============================================================
# ダッシュボードシート（KPIカード＋4グラフ）
# ============================================================
def _fill_rect(ws, c1, r1, c2, r2, color):
    f = T.fill(color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = f


def _kpi_card(ws, col, row, span, label, formula, value_color, fmt, card_color="card"):
    """ラベル(上)＋値(下・大)の2段カードを描く。"""
    c2 = col + span - 1
    _fill_rect(ws, col, row, c2, row + 3, card_color)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=c2)
    lc = ws.cell(row=row, column=col)
    lc.value = label
    lc.font = T.font(size=10, color="subtext")
    lc.alignment = T.LEFT
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 3, end_column=c2)
    vc = ws.cell(row=row + 1, column=col)
    vc.value = formula
    vc.font = T.font(size=20, bold=True, color=value_color)
    vc.alignment = T.LEFT
    vc.number_format = fmt


def build_dashboard(wb):
    ws = wb.create_sheet(SH["dash"])
    hide_grid(ws)
    paint(ws, 60, 16)
    ws.sheet_properties.tabColor = T.COL["accent"]
    ws.sheet_view.showRowColHeaders = False
    title_block(ws, "資産ダッシュボード", "対象年月: （設定シートで変更）")
    put(ws, "B3", '=" 対象年月 " & 対象年月', font_kw=dict(size=10, color="subtext"), align=T.LEFT)

    # 列幅（カード用）
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 9
    for gap in ("E", "I", "M"):
        ws.column_dimensions[gap].width = 2
    ws.column_dimensions["A"].width = 2

    kpi = 'SUMIFS(T_Txn[金額],T_Txn[種類],"{k}",T_Txn[年月],対象年月)'
    income = kpi.format(k="収入")
    expense = kpi.format(k="支出")

    # 1段目 KPI（総資産／今月収入／今月支出）
    _kpi_card(ws, 2, 5, 3, "総資産", "=SUM(T_Account[現在残高])", "accent", T.FMT_YEN0, "card2")
    _kpi_card(ws, 6, 5, 3, "今月収入", f"={income}", "green", T.FMT_YEN0)
    _kpi_card(ws, 10, 5, 3, "今月支出", f"={expense}", "red", T.FMT_YEN0)
    # 2段目 KPI（今月収支／貯蓄率／固定費）
    _kpi_card(ws, 2, 10, 3, "今月収支", f"={income}-{expense}", "text", T.FMT_YEN, "card2")
    _kpi_card(ws, 6, 10, 3, "今月の貯蓄率", f'=IFERROR(({income}-{expense})/{income},"")',
              "yellow", T.FMT_PCT)
    _kpi_card(ws, 10, 10, 3, "固定支出/月",
              '=SUMIFS(T_Fixed[金額],T_Fixed[種類],"支出",T_Fixed[有効],"ON")', "peach", T.FMT_YEN0)
    # 3段目 口座別残高（表示順で最大3口座を動的表示）
    put(ws, "B15", "■ 口座別残高", font_kw=dict(bold=True, color="teal"), align=T.LEFT)
    for i in range(3):
        col = 2 + i * 4
        nm = f'=IFERROR(INDEX(T_Account[口座名],MATCH({i+1},T_Account[表示順],0)),"")'
        bal = f'=IFERROR(INDEX(T_Account[現在残高],MATCH({i+1},T_Account[表示順],0)),"")'
        # ラベル行に口座名（式）、値行に残高（式）を配置
        _kpi_card(ws, col, 16, 3, "", bal, "teal", T.FMT_YEN0)
        ws.cell(row=16, column=col).value = nm

    return ws


# ============================================================
# グラフ生成（openpyxl でデータ参照を作成。視覚のダーク化は COM 段階で実施）
# ============================================================
def build_charts(wb):
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference

    dash = wb[SH["dash"]]
    master = wb[SH["master"]]
    report = wb[SH["report"]]
    snap = wb[SH["snap"]]

    mfirst, mlast = report._rng["month"]
    cfirst, clast = report._rng["cat"]

    CH_H, CH_W = 7.6, 8.6

    # 資産割合（円）: マスタ 口座名(B6:B8) × 現在残高(E6:E8)
    pie = PieChart()
    pie.title = "資産割合"
    pie.height, pie.width = CH_H, CH_W
    pie.add_data(Reference(master, min_col=5, min_row=6, max_row=8), titles_from_data=False)
    pie.set_categories(Reference(master, min_col=2, min_row=6, max_row=8))
    dash.add_chart(pie, "B21")

    # カテゴリ別支出（横棒）: レポート カテゴリ × 当月支出
    barc = BarChart()
    barc.type = "bar"
    barc.title = "カテゴリ別支出（当月）"
    barc.height, barc.width = CH_H, CH_W
    barc.legend = None
    barc.add_data(Reference(report, min_col=3, min_row=cfirst, max_row=clast), titles_from_data=False)
    barc.set_categories(Reference(report, min_col=2, min_row=cfirst, max_row=clast))
    dash.add_chart(barc, "H21")

    # 月別収支（縦棒・収入/支出）: レポート 年月 × 収入,支出
    barm = BarChart()
    barm.type = "col"
    barm.title = "月別収支"
    barm.height, barm.width = CH_H, CH_W
    barm.add_data(Reference(report, min_col=3, max_col=4, min_row=mfirst - 1, max_row=mlast),
                  titles_from_data=True)
    barm.set_categories(Reference(report, min_col=2, min_row=mfirst, max_row=mlast))
    dash.add_chart(barm, "B38")

    # 資産推移（折れ線）: スナップ 年月 × 総資産
    line = LineChart()
    line.title = "資産推移"
    line.height, line.width = CH_H, CH_W
    line.legend = None
    snap_last = 7  # 初期はseed1行のみ（VBAで増える）
    line.add_data(Reference(snap, min_col=3, min_row=7, max_row=snap_last), titles_from_data=False)
    line.set_categories(Reference(snap, min_col=2, min_row=7, max_row=snap_last))
    dash.add_chart(line, "H38")


# ============================================================
# メイン
# ============================================================
def build():
    wb = Workbook()
    # 既定シートは最後に順序調整するため一旦保持
    default = wb.active

    build_txn(wb)
    build_fixed(wb)
    build_sub(wb)
    build_snap(wb)
    build_master(wb)
    build_config(wb)
    build_report(wb)      # 集計（マスタ名・スナップ参照）
    build_dashboard(wb)   # KPIカード（全シート後）
    build_charts(wb)      # グラフ（全データ範囲確定後）

    # 既定シートを削除（実シートを作り終えてから）
    wb.remove(default)

    # シート表示順（左から）: ダッシュボード→取引→固定→サブスク→レポート→スナップ→マスタ→設定
    order = [SH["dash"], SH["txn"], SH["fixed"], SH["sub"],
             SH["report"], SH["snap"], SH["master"], SH["config"]]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    wb.save(OUT_XLSX)
    print(f"[build_structure] saved: {OUT_XLSX}")
    print(f"[build_structure] sheets: {wb.sheetnames}")


if __name__ == "__main__":
    build()
